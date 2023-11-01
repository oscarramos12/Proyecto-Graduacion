import os 
import openpyxl
import fitz
import re
from langdetect import detect


class Color:
    RESET = "\033[0m"
    RED = "\033[91m"
    GREEN = "\033[92m"
    YELLOW = "\033[93m"
    BLUE = "\033[94m"
    MAGENTA = "\033[95m"
    CYAN = "\033[96m"
    WHITE = "\033[97m"


def get_peeks(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        text = file.read()
        if(len(text) > 1500):
            text_portion = 1500
            mid_text = int((len(text)/2))
            bottom_text = len(text) - text_portion

            top_peek = ''
            mid_peek = ''
            bottom_peek = ''

            top_peek = text[0:text_portion]
            mid_peek = text[mid_text:mid_text + text_portion]
            bottom_peek = text[bottom_text: len(text)]
        else:
            top_peek = text
            mid_peek = text
            bottom_peek = text

        return top_peek,mid_peek,bottom_peek
                

def check_language(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            text = file.read()
            language = detect(text)
            return language
    except Exception as e:
        print(f"Error abriendo el archivo o detectando idioma: {e}, {file_path}")
        return None

def check_unique_chars(file_path):

    if os.path.isfile(file_path) and file_path.endswith('.txt'):
        with open(file_path, 'r', encoding='utf-8') as file:
            unique_characters = []
            unique_letters = []
            no_dups = set()
            for line in file:
                words = re.findall(r'\w+', line)

                for i, word in enumerate(words):
                    for char in word:
                        if char.lower() not in 'abcdefghijklmnñopqrstuvwxyzáéíóúü1234567890':
                            if(char not in no_dups):
                                no_dups.add(char)
                                context_words = words[max(0, i - 3):i] + [word] + words[i + 1:i + 4]
                                full_conext =char + ': ' + str(context_words)
                                unique_letters.append(full_conext)
                for check_char in line:
                    if((check_char not in '?¿!¡._-:;"&() ,1234567890\n\'') and (check_char not in no_dups) and not(check_char.isalpha())):
                        no_dups.add(check_char)
                        unique_characters.append(check_char)

        return unique_characters,unique_letters

def extract_font_info(pdf_file):
    font_sizes_count = {}
    font_types_count = {}

    pdf_document = fitz.open(pdf_file)

    for page_number in range(pdf_document.page_count):
        page = pdf_document.load_page(page_number)

        for block in page.get_text("dict")["blocks"]:
            if "lines" in block:
                for line in block["lines"]:
                    if "spans" in line:
                        for span in line["spans"]:
                            font_name = span.get("font", "Unknown")
                            font_size = span.get("size", 0)
                            #text = span.get("text", "")

                            if font_size in font_sizes_count:
                                font_sizes_count[font_size] += 1
                            else:
                                font_sizes_count[font_size] = 1

                            if font_name in font_types_count:
                                font_types_count[font_name] += 1
                            else:
                                font_types_count[font_name] = 1

                            #print(f"Page: {page_number}, Font: {font_name}, Size: {font_size} - Text: {text} ")
    pdf_document.close()
    font_sizes_count = dict(sorted(font_sizes_count.items(), key=lambda item: item[1], reverse=True))
    font_types_count = dict(sorted(font_types_count.items(), key=lambda item: item[1], reverse=True))
    return font_sizes_count, font_types_count

def write_report(report_folder_name,report_file_name,report_sheet_name,text_to_add):
    current_file_path = os.path.abspath(__file__)
    current_directory =  os.path.dirname(current_file_path)
    current_directory = current_directory.replace('scraper.py', '')
    prefix = current_directory +'\\'+ report_folder_name + '\\'

    if not os.path.exists(prefix):
        os.makedirs(prefix)
        print(f"{Color.MAGENTA}Se creo el folder {report_folder_name}{Color.RESET}")

    file_path = prefix + report_file_name  + '.xlsx'

    if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            report_sheet = workbook[report_sheet_name]

            column_values = []
            row_values = []
            variable_max_col = 0

            #si no se estan agregando bien aqui esta el problema

            if(report_sheet_name == 'Archivos Duplicados'):
                variable_max_col = 4
            elif(report_sheet_name == 'Errores Descargas'):
                variable_max_col = 1
            elif(report_sheet_name == 'Contenidos Similares'):
                variable_max_col = 3
            elif(report_sheet_name == 'Caracteristicas Docs'):
                variable_max_col = 3
            elif(report_sheet_name == 'Idiomas Diferentes'):
                variable_max_col = 2 
            elif(report_sheet_name == 'Archivos Descargados'):
                variable_max_col = 2
            elif(report_sheet_name == 'Todos los Caracteres'):
                variable_max_col = 1
            elif(report_sheet_name == 'Links Recorridos'):
                variable_max_col = 1
            elif(report_sheet_name == 'Archivos Vacios'):
                variable_max_col = 1
            elif(report_sheet_name == 'N-Gramas'):
                variable_max_col = 2
            elif(report_sheet_name == 'Modelacion Temas'):
                variable_max_col = 2

            
            for row in report_sheet.iter_rows(min_col=0,max_col=variable_max_col):
                column_values = []
                for cell in row:
                    column_values.append(cell.value)
                row_values.append(column_values)

            if(report_sheet_name == 'Archivos Duplicados'):
                if(text_to_add not in row_values):
                    report_sheet.append([text_to_add[0],text_to_add[1],text_to_add[2],text_to_add[3]])
            elif(report_sheet_name == 'Errores Descargas'):
                 if(text_to_add not in row_values):
                    report_sheet.append([text_to_add[0],text_to_add[1]])
            elif(report_sheet_name == 'Contenidos Similares'):
                 if(text_to_add not in row_values):
                    report_sheet.append([text_to_add[0],text_to_add[1],text_to_add[2]])
            elif(report_sheet_name == 'Idiomas Diferentes'):
                 if(text_to_add not in row_values):
                    report_sheet.append([text_to_add[0],text_to_add[1]])
            elif(report_sheet_name == 'Archivos Descargados'):
                 if(text_to_add not in row_values):
                    report_sheet.append([text_to_add[0],text_to_add[1]])
            elif(report_sheet_name == 'Todos los Caracteres'):
                 if(text_to_add not in row_values):
                    report_sheet.append([text_to_add[0]])
            elif(report_sheet_name == 'Links Recorridos'):
                 if(text_to_add not in row_values):
                    report_sheet.append([text_to_add[0]])
            elif(report_sheet_name == 'Archivos Vacios'):
                 if(text_to_add not in row_values):
                    report_sheet.append([text_to_add[0]])
            elif(report_sheet_name == 'Caracteristicas Docs'):
                 if(text_to_add[0:3] not in row_values):
                    if(len(text_to_add) == 9):
                        report_sheet.append([text_to_add[0],text_to_add[1],text_to_add[2],text_to_add[3],text_to_add[4],text_to_add[5],text_to_add[6],text_to_add[7],text_to_add[8]])
                    else:
                        report_sheet.append([text_to_add[0],text_to_add[1],text_to_add[2],text_to_add[3],text_to_add[4],text_to_add[5],text_to_add[6]])
            elif(report_sheet_name == 'N-Gramas'):
                 if(text_to_add not in row_values):
                    report_sheet.append([text_to_add[0],text_to_add[1],text_to_add[2],text_to_add[3],text_to_add[4],text_to_add[5]])
            elif(report_sheet_name == 'Modelacion Temas'):
                 if(text_to_add not in row_values):
                    report_sheet.append([text_to_add[0],text_to_add[1],text_to_add[2]])

            workbook.save(file_path)
    elif(report_file_name == 'general_reports'):
        workbook = openpyxl.Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        sheet_names = ['Links Recorridos','Archivos Duplicados','Errores Descargas','Contenidos Similares','Idiomas Diferentes','Modelacion Temas', 'Archivos Descargados', 'Todos los Caracteres','Archivos Vacios']
        for sheet_name in sheet_names:
            workbook.create_sheet(title=sheet_name)
        
        duplicate_sheet = workbook['Archivos Duplicados']
        duplicate_sheet.append(['Archivo a Descargar','Archivo Original','Score','Link de Descarga'])
        if(report_sheet_name == 'Archivos Duplicados'):
            duplicate_sheet.append([text_to_add[0],text_to_add[1],text_to_add[2],text_to_add[3]])

        error_sheet = workbook['Errores Descargas']
        error_sheet.append(['Link de Descarga','Codigo Error'])
        if(report_sheet_name == 'Errores Descargas'):
            error_sheet.append([text_to_add[0],text_to_add[1]])

        contents_sheet = workbook['Contenidos Similares']
        contents_sheet.append(['Titulo Libro 1', 'Titulo Libro 2', 'Similitud'])
        if(report_sheet_name == 'Contenidos Similares'):
            contents_sheet.append([text_to_add[0],text_to_add[1],text_to_add[2]])

        contents_sheet = workbook['Idiomas Diferentes']
        contents_sheet.append(['Titulo Archivo', 'Idioma'])
        if(report_sheet_name == 'Idiomas Diferentes'):
            contents_sheet.append([text_to_add[0],text_to_add[1]])

        contents_sheet = workbook['Archivos Descargados']
        contents_sheet.append(['Titulo Archivo','Link de Descarga'])
        if(report_sheet_name == 'Archivos Descargados'):
            contents_sheet.append([text_to_add[0],text_to_add[1]])

        contents_sheet = workbook['Todos los Caracteres']
        contents_sheet.append(['Caracteres'])
        if(report_sheet_name == 'Todos los Caracteres'):
            contents_sheet.append([text_to_add[0]])

        contents_sheet = workbook['Links Recorridos']
        contents_sheet.append(['Link'])
        if(report_sheet_name == 'Links Recorridos'):
            contents_sheet.append([text_to_add[0]])

        contents_sheet = workbook['Archivos Vacios']
        contents_sheet.append(['Nombre Archivo'])
        if(report_sheet_name == 'Archivos Vacios'):
            contents_sheet.append([text_to_add[0]])

        contents_sheet = workbook['Modelacion Temas']
        contents_sheet.append(['Nombre Archivo','Palabras','Valor Relevancia'])
        if(report_sheet_name == 'Modelacion Temas'):
            contents_sheet.append([text_to_add[0],text_to_add[1],text_to_add[2]])
    
        workbook.save(file_path)

    elif(report_file_name == 'dirty_files_reports'):
        workbook = openpyxl.Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        workbook.create_sheet(title='Caracteristicas Docs')
        workbook.create_sheet(title='N-Gramas')
        workbook.create_sheet(title='Modelacion Temas')

        contents_sheet = workbook['Caracteristicas Docs']
        contents_sheet.append(['Titulo Documento', 'Fonts Documento', 'Tamaño Fonts','Caracteres Unicos','Letras Unicas','Idioma','Top Peek', 'Mid Peek', 'Bottom Peek'])
        if(report_sheet_name == 'Caracteristicas Docs'):
            contents_sheet.append([text_to_add[0],text_to_add[1],text_to_add[2],text_to_add[3],text_to_add[4],text_to_add[5],text_to_add[6],text_to_add[7],text_to_add[8]])

        contents_sheet = workbook['N-Gramas']
        contents_sheet.append(['Nombre Archivo','Tipo N-Grama','Top','Veces','Bottom','Veces'])
        if(report_sheet_name == 'N-Gramas'):          
            contents_sheet.append([text_to_add[0],text_to_add[1],text_to_add[2],text_to_add[3],text_to_add[4],text_to_add[5]])

        contents_sheet = workbook['Modelacion Temas']
        contents_sheet.append(['Nombre Archivo','Palabras','Valor Relevancia'])
        if(report_sheet_name == 'Modelacion Temas'):    
            contents_sheet.append([text_to_add[0],text_to_add[1],text_to_add[2]])


        workbook.save(file_path)
        
        print(f"{Color.MAGENTA} Se creo el archivo de reporte {report_file_name} {Color.RESET}")

    elif(report_file_name == 'clean_files_reports'):
        workbook = openpyxl.Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        workbook.create_sheet(title='Caracteristicas Docs')
        workbook.create_sheet(title='N-Gramas')
        workbook.create_sheet(title='Modelacion Temas')

        contents_sheet = workbook['Caracteristicas Docs']
        contents_sheet.append(['Titulo Documento','Caracteres Unicos','Letras Unicas','Idioma','Top Peek', 'Mid Peek', 'Bottom Peek'])
        if(report_sheet_name == 'Caracteristicas Docs'):
            contents_sheet.append([text_to_add[0],text_to_add[1],text_to_add[2],text_to_add[3],text_to_add[4],text_to_add[5],text_to_add[6]])

        contents_sheet = workbook['N-Gramas']
        contents_sheet.append(['Nombre Archivo','Tipo N-Grama','Top','Veces','Bottom','Veces'])
        if(report_sheet_name == 'N-Gramas'):          
            contents_sheet.append([text_to_add[0],text_to_add[1],text_to_add[2],text_to_add[3],text_to_add[4],text_to_add[5]])

        contents_sheet = workbook['Modelacion Temas']
        contents_sheet.append(['Nombre Archivo','Palabras','Valor Relevancia'])
        if(report_sheet_name == 'Modelacion Temas'):    
            contents_sheet.append([text_to_add[0],text_to_add[1],text_to_add[2]])


        workbook.save(file_path)
        
        print(f"{Color.MAGENTA} Se creo el archivo de reporte {report_file_name} {Color.RESET}")

        
def generate_clean_file_reports(folder_path):
    if not os.path.exists(folder_path):
        print(f"El folder no existe: [{folder_path}]")
    else:

        for file_name in os.listdir(folder_path):
            txt_file_path = os.path.join(folder_path, file_name)
            clean_file_name = file_name.replace('.txt','')


            unique_characters,unique_letters = check_unique_chars(txt_file_path)
            language = check_language(txt_file_path)
            top_peek, mid_peek, bot_peek = get_peeks(txt_file_path)

            report_data = []

            report_data.append(clean_file_name)
            report_data.append(str(unique_characters))
            report_data.append(str(unique_letters))
            report_data.append(language)
            report_data.append(top_peek)
            report_data.append(mid_peek)
            report_data.append(bot_peek)

            write_report('file_reports','clean_files_reports','Caracteristicas Docs',report_data)

def generate_dirty_file_reports(txt_folder_path,pdf_folder_path):

    if not os.path.exists(txt_folder_path):
        print(f"El folder no existe: [{txt_folder_path}]")
    else:
        pdf_file_names = []
        full_names = os.listdir(pdf_folder_path)
        for name in full_names:
            pdf_file_names.append(name.replace('.pdf',''))

        for file_name in os.listdir(txt_folder_path):
            txt_file_path = os.path.join(txt_folder_path, file_name)
            clean_file_name = file_name.replace('.txt','')

            if(clean_file_name in pdf_file_names):

                sizes_count, types_count = extract_font_info(pdf_folder_path + clean_file_name + '.pdf')
                unique_characters,unique_letters = check_unique_chars(txt_file_path)
                language = check_language(txt_file_path)
                top_peek, mid_peek, bot_peek = get_peeks(txt_file_path)

                report_data = []

                report_data.append(clean_file_name)
                report_data.append(str(types_count))
                report_data.append(str(sizes_count))
                report_data.append(str(unique_characters))
                report_data.append(str(unique_letters))
                report_data.append(language)
                report_data.append(top_peek)
                report_data.append(mid_peek)
                report_data.append(bot_peek)

                write_report('file_reports','dirty_files_reports','Caracteristicas Docs',report_data)

            
folder_path = 'C:\\Users\\Oscar\\Desktop\\LIBROS\\downloaded_pdfs\\'
txt_path = 'C:\\Users\\Oscar\\Desktop\\LIBROS\\txt_files\\'
clean_txt_path = 'C:\\Users\\Oscar\\Desktop\\LIBROS\\txt_clean\\'
#generate_dirty_file_reports(txt_path,folder_path)
#generate_clean_file_reports(clean_txt_path)